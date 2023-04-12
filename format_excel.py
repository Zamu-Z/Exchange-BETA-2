import pathlib
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from page1 import *
from page2 import *
import pandas as pd
from website import *

column = ['A', 'B', 'C', 'D', 'E']
title_1 = ["ประเทศ", "สกุลเงิน", "ตั๋วเงิน", "เงินโอน", "อัตราขายถัวเฉลี่ย"]
title_2 = ["ประเทศ", "สกุลเงิน", "อัตราซื้อ", "อัตราขาย"]


def dataframe_data(name):
    df1 = pd.DataFrame(data=data_1(), columns=title_1)
    df2 = pd.DataFrame(data=data_2(), columns=title_2)
    with pd.ExcelWriter(name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as book:
        df1.to_excel(book, sheet_name="Page1", startcol=0, startrow=7, index=False, header=False)
        df2.to_excel(book, sheet_name="Page2", startcol=0, startrow=4, index=False, header=True)


def format_excel(name):
    file = pathlib.Path(name)
    if file.exists() is True:
        book = openpyxl.load_workbook(name)
        ws1 = book[book.sheetnames[0]]
        ws2 = book[book.sheetnames[1]]

        ws1['A1'].value = title1()
        ws1['A2'].value = title_page1()
        ws1['D4'].value = "".join(description_1()[2])
        ws1['C6'].value = "อัตราซื้อถัวเฉลี่ย"
        for t, ti in enumerate(column):
            if ti == "C" or ti == "D":
                ws1[ti + "7"].value = title_1[t]
            else:
                ws1[ti + "6"].value = title_1[t]
        ws2["A1"].value = title1()
        book.save(name)

        dataframe_data(name)
    else:
        book = openpyxl.Workbook()
        book.active.title = "Page1"
        ws1 = book["Page1"]
        ws2 = book.create_sheet("Page2", 1)
        font = Font(bold=True)
        align = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(fill_type="solid", start_color="89eb34")
        border = Border(left=Side(border_style='thin', color="000000"),
                        right=Side(border_style='thin', color="000000"),
                        top=Side(border_style='thin', color="000000"),
                        bottom=Side(border_style='thin', color="000000"))

        # sheet 1
        ws1.merge_cells("A1:E1"), ws1.merge_cells("A2:E2"), ws1.merge_cells("A29:E29")
        ws1.merge_cells("A4:C4"), ws1.merge_cells("D4:E4")

        ws1.merge_cells("A6:A7"), ws1.merge_cells("B6:B7"), ws1.merge_cells("E6:E7")
        ws1.merge_cells("C6:D6")

        for n, c in enumerate(column):
            if c == "A" or c == "E":
                ws1.column_dimensions[c].width = 22
                ws2.column_dimensions[c].width = 22
            else:
                ws1.column_dimensions[c].width = 16
                ws2.column_dimensions[c].width = 16

        ws1['A1'].font = font
        ws1['A1'].alignment = align
        ws1['A1'].value = title1()

        ws1['A2'].font = font
        ws1['A2'].alignment = align
        ws1['A2'].value = title_page1()

        ws1['A4'].value = "".join(description_1()[0:2])
        ws1['D4'].value = "".join(description_1()[2])

        ws1['C6'].value = "อัตราซื้อถัวเฉลี่ย"

        ws1['A6'].alignment = align

        for n, t in enumerate(column):
            ws1[t + "6"].alignment = align
            ws1[t + "6"].fill = fill
            ws1[t + "6"].border = border
            ws1[t + "7"].border = border
            if t == "C" or t == "D":
                ws1[t + "7"].alignment = align
                ws1[t + "7"].fill = fill

        for c in column:
            for num in range(8, 27):
                ws1[c + str(num)].border = border
                if num == 26:
                    ws1[c + str(num)].border = border

        for b in range(8, 27):
            b_row = 'B' + str(b)
            ws1[b_row].alignment = align

        for t, ti in enumerate(column):
            if ti == "C" or ti == "D":
                ws1[ti + "7"].value = title_1[t]
            else:
                ws1[ti + "6"].value = title_1[t]

        # sheet 2
        ws2.merge_cells("A1:D1"), ws2.merge_cells("A3:E3")
        ws2["A1"].value = title1()
        ws2["A1"].font = font
        ws2["A1"].alignment = align
        ws2["A3"].value = title2(1)

        for c in column[0:4]:
            ws2[c + "5"].fill = fill
            ws2[c + "5"].border = border
            for num_col in range(6, 35):
                ws2["B" + str(num_col)].alignment = align
                ws2[c + str(num_col)].border = border
                if num_col == 34:
                    ws2[c + str(num_col)].border = border

        ws2.merge_cells("A36:E36"), ws2.merge_cells("A38:K38")

        ws1.sheet_view.showGridLines = False
        ws2.sheet_view.showGridLines = False

        book.save(name)

        dataframe_data(name)


if __name__ == '__main__':
    print(process_time())
